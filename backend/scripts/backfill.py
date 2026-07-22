"""Backfill vehicles + price_history from the legacy monthly Google Sheets
workbooks (exported as .xlsx) into the app's database.

Handles the two structurally different layouts confirmed during requirements
gathering:
  - "used" workbooks (Chevy Used, Subaru Used, SARs): each tab is a flat list
    of vehicle rows, sometimes with more than one Stock#-headered sub-block
    per tab (e.g. Internet tab = "Sheet #1" + "Sheet #2").
  - "new" workbooks (Chevy New, Subaru New): each tab (Car/SUV/Truck) contains
    one stacked block per model (Colorado, Silverado 1500, ...).

Run from backend/: `python -m scripts.backfill`
"""

import datetime
import re
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.db import Base, SessionLocal, engine  # noqa: E402
from app.models import PriceHistory, Store, UnwindEvent, StoreTransfer, Vehicle  # noqa: E402


def _clear_existing(db, store_id: int, store_type: str):
    """Bulk delete bypasses ORM cascade, so clear child rows explicitly."""
    vehicle_ids = [
        v.id
        for v in db.query(Vehicle.id).filter(
            Vehicle.current_store_id == store_id, Vehicle.store_type == store_type
        )
    ]
    if not vehicle_ids:
        return
    db.query(PriceHistory).filter(PriceHistory.vehicle_id.in_(vehicle_ids)).delete(synchronize_session=False)
    db.query(UnwindEvent).filter(UnwindEvent.vehicle_id.in_(vehicle_ids)).delete(synchronize_session=False)
    db.query(StoreTransfer).filter(StoreTransfer.vehicle_id.in_(vehicle_ids)).delete(synchronize_session=False)
    db.query(Vehicle).filter(Vehicle.id.in_(vehicle_ids)).delete(synchronize_session=False)
    db.commit()

SNAPSHOT_DIR = Path(r"C:\Users\lv-us\Projects\Snapshot")

USED_FILES = {
    "chevy": "JULY  CHEVY  USED CAR SNAP SHOT 2026.xlsx",
    "subaru": "TFS JULY SNAPSHOT 2026.xlsx",
    "sars": "JULY  VALUE SaRsLOT   USED CAR SNAP SHOT 2025.xlsx",
}
NEW_FILES = {
    "chevy": "JULY NEW CAR SNAP 2026.xlsx",
    "subaru": "NC Snapshot - 07_26.xlsx",
}

USED_SKIP_TABS = {"Snapshot", "Strategy", "ALL ON ONE"}
NEW_SKIP_TABS = {"Snapshot", "Strategy"}

# normalized-header -> our field name (used-tab layout)
HEADER_MAP = {
    "stock #": "stock_number",
    "sold": "_sold_flag",
    "real age": "_real_age",
    "status": "status",
    "price rank": "price_rank",
    "vehicle type": "vehicle_type",
    "age 1st of month": "inventory_date",
    "date in inventory": "inventory_date",
    "date": "inventory_date",
    "price": "price",
    "cost": "cost",
    "available gross": "_available_gross",
    "sold gross": "_sold_gross",
    "sold age": "_sold_age",
    "sold price": "sold_price",
    "sold cost": "sold_cost",
    "source": "source",
    "trade evaluator": "trade_evaluator",
}

YMM_RE = re.compile(r"^\s*(\d{2})\s+([A-Za-z.]+)\s+(.*)$")


def year_from_2digit(yy: int) -> int:
    return 2000 + yy if yy <= 40 else 1900 + yy


def parse_ymm(vehicle_type: str):
    """Best-effort split of the free-text 'Vehicle Type' blob into year/make/model/trim.

    No VIN-decode is available in the source sheets, so this is regex-based;
    per the client, some rows will need manual cleanup post-import.
    """
    if not vehicle_type:
        return None, None, None, vehicle_type, None
    m = YMM_RE.match(vehicle_type.strip())
    if not m:
        return None, None, None, vehicle_type.strip(), None
    yy, make, rest = m.groups()
    parts = rest.split(None, 1)
    model = parts[0] if parts else None
    trim = parts[1] if len(parts) > 1 else None
    mileage = None
    mm = re.search(r"([\d,]+)\s*K\b", rest, re.IGNORECASE)
    if mm:
        try:
            mileage = int(mm.group(1).replace(",", "")) * 1000
        except ValueError:
            mileage = None
    return year_from_2digit(int(yy)), make.title(), model, trim, mileage


def normalize_header(v):
    return str(v).strip().lower() if v is not None else None


def to_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def parse_used_tab(ws):
    """Yields dict rows for every Stock#-headered sub-block in a used-type tab."""
    header_rows = []
    for r in range(1, ws.max_row + 1):
        if normalize_header(ws.cell(row=r, column=1).value) == "stock #":
            header_rows.append(r)

    for idx, header_row in enumerate(header_rows):
        col_map = {}
        for c in range(1, ws.max_column + 1):
            key = normalize_header(ws.cell(row=header_row, column=c).value)
            if key in HEADER_MAP:
                col_map[HEADER_MAP[key]] = c

        next_header = header_rows[idx + 1] if idx + 1 < len(header_rows) else ws.max_row + 1
        blank_streak = 0
        r = header_row + 1
        while r < next_header and blank_streak < 3:
            stock = ws.cell(row=r, column=col_map.get("stock_number", 1)).value
            status_val = ws.cell(row=r, column=col_map.get("status", 4)).value
            b_col2 = ws.cell(row=r, column=2).value
            if isinstance(b_col2, str) and "TOTAL" in b_col2.upper():
                r += 1
                continue
            # Real stock numbers are always alphanumeric strings. Some tabs (e.g.
            # RENTALS) have an unrelated second sub-table further down — like a
            # VIN/registration list — whose first column holds a year or date
            # instead. Treat any non-string value here as end-of-block, not data.
            if not stock or not isinstance(stock, str) or not stock.strip():
                blank_streak += 1
                r += 1
                continue
            blank_streak = 0

            row = {"stock_number": str(stock).strip(), "status": status_val}
            for field, col in col_map.items():
                if field in ("stock_number", "status"):
                    continue
                row[field] = ws.cell(row=r, column=col).value
            yield row
            r += 1


def parse_new_tab(ws):
    """Yields dict rows for each model-block in a new-type tab (Car/SUV/Truck)."""
    header_rows = []
    for r in range(1, ws.max_row + 1):
        if normalize_header(ws.cell(row=r, column=1).value) == "stock #":
            header_rows.append(r)

    for idx, header_row in enumerate(header_rows):
        model_name = None
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row - 1, column=c).value
            if isinstance(val, str) and val.strip():
                model_name = val.strip()
                break
        col_map = {}
        for c in range(1, ws.max_column + 1):
            key = normalize_header(ws.cell(row=header_row, column=c).value)
            if key == "stock #":
                col_map["stock_number"] = c
            elif key == "status":
                col_map["status"] = c
            elif key in ("trim / pkg",):
                col_map["trim_details"] = c
            elif key == "color":
                col_map["color"] = c
            elif key == "msrp":
                col_map["price"] = c
            elif key == "cost":
                col_map["cost"] = c
            elif key == "selling price":
                col_map["sold_price"] = c
            elif key in ("date", "age 1st of month"):
                col_map["inventory_date"] = c

        next_header = header_rows[idx + 1] if idx + 1 < len(header_rows) else ws.max_row + 1
        blank_streak = 0
        r = header_row + 1
        while r < next_header and blank_streak < 4:
            stock = ws.cell(row=r, column=col_map.get("stock_number", 1)).value
            # Real stock numbers are always alphanumeric strings. Some tabs (e.g.
            # RENTALS) have an unrelated second sub-table further down — like a
            # VIN/registration list — whose first column holds a year or date
            # instead. Treat any non-string value here as end-of-block, not data.
            if not stock or not isinstance(stock, str) or not stock.strip():
                blank_streak += 1
                r += 1
                continue
            blank_streak = 0
            row = {"stock_number": str(stock).strip(), "model_hint": model_name}
            for field, col in col_map.items():
                if field == "stock_number":
                    continue
                row[field] = ws.cell(row=r, column=col).value
            yield row
            r += 1


def build_vehicle(row, store: Store, store_type: str, bucket: str):
    v = Vehicle(
        stock_number=row.get("stock_number"),
        originating_store_id=store.id,
        current_store_id=store.id,
        store_type=store_type,
        bucket=bucket,
        status=str(row.get("status")).strip() if row.get("status") else "LOT",
        price=row.get("price") if isinstance(row.get("price"), (int, float)) else None,
        cost=row.get("cost") if isinstance(row.get("cost"), (int, float)) else None,
        sold_price=row.get("sold_price") if isinstance(row.get("sold_price"), (int, float)) else None,
        sold_cost=row.get("sold_cost") if isinstance(row.get("sold_cost"), (int, float)) else None,
        inventory_date=to_date(row.get("inventory_date")),
        price_rank=int(row["price_rank"]) if isinstance(row.get("price_rank"), (int, float)) else None,
    )

    if store_type == "used" and row.get("vehicle_type"):
        year, make, model, trim, mileage = parse_ymm(str(row["vehicle_type"]))
        v.year, v.make, v.model, v.trim_details, v.mileage = year, make, model, trim, mileage
    elif store_type == "new":
        v.model = row.get("model_hint")
        v.trim_details = row.get("trim_details")
        v.year = 2026  # source files are current-model-year only; no explicit year column

    return v


def run():
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    stores = {s.slug: s for s in db.query(Store).all()}
    if not stores:
        raise SystemExit("No stores found — run `python -m app.seed` first.")

    stats = {}

    for slug, filename in USED_FILES.items():
        store = stores[slug]
        # Idempotent: clear any prior import for this exact store/type before
        # reloading, so re-running this script can never create duplicates.
        _clear_existing(db, store.id, "used")
        wb = openpyxl.load_workbook(SNAPSHOT_DIR / filename, data_only=True)
        count, gross_sum, sold_count = 0, 0.0, 0
        for tab_name in wb.sheetnames:
            if tab_name in USED_SKIP_TABS:
                continue
            ws = wb[tab_name]
            for row in parse_used_tab(ws):
                v = build_vehicle(row, store, "used", bucket=tab_name)
                db.add(v)
                count += 1
                if v.price is not None and v.cost is not None:
                    gross_sum += v.price - v.cost
                if v.status and v.status.upper() == "SOLD":
                    sold_count += 1
                    if v.sold_price is not None:
                        db.flush()
                        db.add(
                            PriceHistory(
                                vehicle_id=v.id, old_price=None, new_price=v.sold_price, changed_by="backfill"
                            )
                        )
        db.commit()
        stats[f"{slug}-used"] = {"imported": count, "available_gross_sum": round(gross_sum, 2), "sold_count": sold_count}

    for slug, filename in NEW_FILES.items():
        store = stores[slug]
        _clear_existing(db, store.id, "new")
        wb = openpyxl.load_workbook(SNAPSHOT_DIR / filename, data_only=True)
        count, gross_sum, sold_count = 0, 0.0, 0
        for tab_name in wb.sheetnames:
            if tab_name in NEW_SKIP_TABS:
                continue
            ws = wb[tab_name]
            for row in parse_new_tab(ws):
                v = build_vehicle(row, store, "new", bucket=tab_name)
                db.add(v)
                count += 1
                if v.price is not None and v.cost is not None:
                    gross_sum += v.price - v.cost
                if v.status and v.status.upper() == "SOLD":
                    sold_count += 1
        db.commit()
        stats[f"{slug}-new"] = {"imported": count, "available_gross_sum": round(gross_sum, 2), "sold_count": sold_count}

    db.close()
    print("Backfill complete:")
    for k, v in stats.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    run()
